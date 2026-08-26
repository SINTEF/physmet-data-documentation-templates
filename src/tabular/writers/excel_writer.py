from pathlib import Path
from typing import Union, Any, Optional

import tabular.models
from .base import BaseWriter


class ExcelWriter(BaseWriter):
    """Writes tabular data to an Excel (.xlsx) workbook."""

    def write(
        self,
        data: Union[tabular.models.Table, tabular.models.Tables],
        path: Optional[Path] = None,
        **kwargs: Any,
    ) -> Optional[str]:
        """
        Writes data to an Excel workbook. It creates one spreadsheet per Table in
        the provided collection.

        Args:
            data (Union[tabular.models.Table, tabular.models.Tables]): The dataset(s) to export.
            path (Optional[Path], optional): Output destination path.
            **kwargs: Reserved for future Excel-specific parameters.

        Returns:
            None (Excel files cannot be parsed as strings in this library).

        Raises:
            ValueError: If path is None, as binary formats cannot be cleanly serialized to standard strings.
            ImportError: If openpyxl is not installed.
        """
        try:
            import openpyxl
        except ImportError as exc:
            raise ImportError(
                "The 'openpyxl' package is required to write Excel files. "
                "Install it using 'pip install openpyxl' or 'pip install PhysMetDatadoc[excel]'."
            ) from exc

        if path is None:
            raise ValueError(
                "Excel format is binary and cannot be generated as a string. You must provide a path."
            )

        self._ensure_directory(path)
        collection = self._ensure_tables(data)
        wb = openpyxl.Workbook()

        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        for i, table in enumerate(collection.tables):
            sheet_title = (table.name or f"Sheet{i}")[:31]
            ws = wb.create_sheet(title=sheet_title)
            ws.append(table.headers)
            for row in table.rows:
                ws.append(row)

        wb.save(path)
        return None
